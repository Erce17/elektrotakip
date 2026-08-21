"""İndirilebilir şablon ve şablonla içe aktarma.

Şablon 5 sütunken (Kategori · Teknik Özellik · Marka · Birim Fiyat · Birim) gerçek
tedarikçi dosyasından **daha zayıf** girdiydi: ham dosyadan kesit, tedarikçi kodu ve
para birimi çıkıyordu, kendi şablonumuzdan çıkmıyordu. Kesit serbest metne çıplak sayı
olarak yazılınca ayrıştırıcı onu okumuyor — ve okumaması doğru, çünkü serbest metinde
`NxM` kalıbı çoğunlukla kablo değil (bkz. `test_product_search.py`).

Bu testler şablonun artık ham dosya kadar bilgi taşıdığını sabitliyor.
"""

import io

import openpyxl
import pytest

from app.excel_import import baslik_satirini_bul, sayfalari_oku
from app.models import Category, Product
from app.routers.catalog import SABLON_SUTUNLARI, kdv_orani_coz


@pytest.fixture
def kullanici(client, make_user, login_as):
    user, token = make_user("sablonci@example.com")
    login_as(token)
    return user


def sablonu_indir(client) -> bytes:
    cevap = client.get("/catalog/download-template")
    assert cevap.status_code == 200
    return cevap.content


def sablonu_doldur(satirlar: list[list]) -> bytes:
    """İndirilen şablonun başlıklarıyla bellekte bir dosya üretir."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ElektroTakip_Sablon"
    ws.append(SABLON_SUTUNLARI)
    for satir in satirlar:
        ws.append(satir)
    akis = io.BytesIO()
    wb.save(akis)
    return akis.getvalue()


def yukle(client, icerik: bytes, dosya_adi="sablon.xlsx"):
    return client.post(
        "/catalog/import-excel",
        files={"file": (dosya_adi, icerik,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )


# =============================================================================
# Şablonun kendisi
# =============================================================================


def test_sablon_indirilebiliyor(client, kullanici):
    icerik = sablonu_indir(client)
    wb = openpyxl.load_workbook(io.BytesIO(icerik))
    assert wb.sheetnames == ["ElektroTakip_Sablon"]
    basliklar = [h.value for h in wb.active[1]]
    assert basliklar == SABLON_SUTUNLARI


@pytest.mark.parametrize(
    "sutun", ["Stok Kodu", "Kesit", "Para Birimi", "KDV %"]
)
def test_sablonda_eksik_olan_sutunlar_var(client, kullanici, sutun):
    """Dördü de sonradan eklendi; olmadan şablon ham dosyadan zayıf kalıyordu."""
    assert sutun in SABLON_SUTUNLARI


def test_sablonun_kendi_basliklari_taninir(client, kullanici):
    """Şablonu üreten yer ile okuyan yer ayrı; ikisi ayrışırsa kimse fark etmez."""
    icerik = sablonu_indir(client)
    (_, satirlar), = list(sayfalari_oku(icerik, "sablon.xlsx"))
    indeks, sutunlar = baslik_satirini_bul(satirlar)

    assert indeks == 0
    assert set(sutunlar) == {
        "kod", "grup", "aciklama", "marka", "kesit", "fiyat", "para_birimi",
        "birim", "kdv",
    }


def test_bos_sablon_yuklenirse_urun_olusmaz(client, db_session, kullanici):
    """Şablonda örnek satır yok; indirip olduğu gibi yüklemek çöp ürün üretmemeli."""
    cevap = yukle(client, sablonu_indir(client))

    assert "eklendi=0" in cevap.headers["location"]
    assert db_session.query(Product).count() == 0


# =============================================================================
# Şablonla içe aktarma
# =============================================================================


def test_sablondaki_kesit_urune_yaziliyor(client, db_session, kullanici):
    """Asıl mesele buydu: kesit kendi sütununda olunca ayrıştırıcı güvenle okuyor."""
    yukle(client, sablonu_doldur([
        ["OZ-001", "Kablo", "H07V-U (NYA) 450/750 V", "Öznur", "3x2.5",
         "41.990", "TL", "KM", "20"],
    ]))

    urun = db_session.query(Product).one()
    assert float(urun.cross_section) == 2.5
    assert urun.core_count == 3


def test_kesit_aciklamaya_gomulurse_okunmaz(client, db_session, kullanici):
    """Erce'nin dün hazırladığı şablonlar böyleydi: '... (CPR Class: Eca) 2.5'.

    Ayrıştırıcı serbest metinde çıplak sayıyı kesit saymıyor ve **saymamalı** —
    Viko'da 532, Klemsan'da 496 satır `NxM` kalıbı içeriyor ve hiçbiri kablo değil.
    Çözüm parser'ı gevşetmek değil, kesiti kendi sütununa yazmak.
    """
    yukle(client, sablonu_doldur([
        ["OZ-002", "Kablo", "H07V-U (NYA) 450/750 V (CPR Class: Eca) 2.5", "Öznur",
         "", "41.990", "TL", "KM", "20"],
    ]))

    assert db_session.query(Product).one().cross_section is None


def test_sablondaki_tedarikci_kodu_yaziliyor(client, db_session, kullanici):
    yukle(client, sablonu_doldur([
        ["OZ-003", "Kablo", "NYA", "Öznur", "16", "100", "TL", "Metre", "20"],
    ]))
    assert db_session.query(Product).one().supplier_code == "OZ-003"


def test_sablondaki_para_birimi_yaziliyor(client, db_session, kullanici):
    """Klemsan tamamen EURO; şablonda para birimi olmasa hepsi TL sanılırdı."""
    yukle(client, sablonu_doldur([
        ["KL-001", "Klemens", "REL 24V DC", "Klemsan", "", "10,58", "EURO", "Adet", "20"],
    ]))
    assert db_session.query(Product).one().currency == "EUR"


def test_para_birimi_bossa_tl_varsayilir(client, db_session, kullanici):
    yukle(client, sablonu_doldur([
        ["X-1", "Kablo", "NYA", "Öznur", "", "100", "", "Metre", "20"],
    ]))
    assert db_session.query(Product).one().currency == "TRY"


def test_sablondaki_kdv_urune_yaziliyor(client, db_session, kullanici):
    """Karışık KDV oranlı tek teklifte tek oran yanlış rakam verir."""
    yukle(client, sablonu_doldur([
        ["A-1", "Malzeme", "yirmilik", "M", "", "100", "TL", "Adet", "20"],
        ["A-2", "Malzeme", "onluk", "M", "", "100", "TL", "Adet", "10"],
        ["A-3", "Malzeme", "birlik", "M", "", "100", "TL", "Adet", "%1"],
    ]))

    oranlar = {u.supplier_code: u.vat_rate for u in db_session.query(Product).all()}
    assert oranlar == {"A-1": 20, "A-2": 10, "A-3": 1}


def test_kdv_bossa_yirmi_varsayilir(client, db_session, kullanici):
    yukle(client, sablonu_doldur([
        ["A-1", "Malzeme", "test", "M", "", "100", "TL", "Adet", ""],
    ]))
    assert db_session.query(Product).one().vat_rate == 20


@pytest.mark.parametrize(
    "ham,beklenen",
    [
        ("20", 20), ("%20", 20), ("20,0", 20), ("10", 10), ("1", 1), ("0", 0),
        ("", 20), ("   ", 20), (None, 20),
        # Kesirli yazım: '0,20' yazan %20 kastediyor. Reddetmek yerine çeviriyoruz —
        # reddetmek '0,10' için sessizce %20 yazardı, yani yanlış rakam.
        ("0,20", 20), ("0,1", 10), ("0,01", 1),
        ("120", 20),       # aralık dışı
        ("-5", 20),        # aralık dışı
        ("bilgi yok", 20),
    ],
)
def test_kdv_orani_okuma(ham, beklenen):
    assert kdv_orani_coz(ham) == beklenen


def test_sablon_km_fiyatini_metreye_ceviriyor(client, db_session, kullanici):
    """Birim sütunu 'KM' diyorsa kutucuk işaretsiz olsa da çevrilir."""
    yukle(client, sablonu_doldur([
        ["OZ-004", "Kablo", "NYA", "Öznur", "2.5", "41.990", "TL", "KM", "20"],
    ]))

    urun = db_session.query(Product).one()
    assert urun.unit == "Metre"
    assert float(urun.unit_price) == pytest.approx(41.99)


def test_sablonla_gelen_urun_aranabiliyor(client, db_session, kullanici):
    """Uçtan uca: şablon → ürün → arama. Bu işin varlık sebebi."""
    from app.product_search import urun_ara

    yukle(client, sablonu_doldur([
        ["OZ-005", "Kablo", "H07V-U (NYA) 450/750 V", "Öznur", "3x2.5",
         "41.990", "TL", "KM", "20"],
        ["OZ-006", "Kablo", "H07V-U (NYA) 450/750 V", "Öznur", "5x16",
         "176.150", "TL", "KM", "20"],
    ]))

    sorgu = db_session.query(Product).join(Category)
    sonuc = urun_ara(sorgu, Product, "3x2.5 NYA")
    assert [u.supplier_code for u in sonuc] == ["OZ-005"]


def test_ayni_kesitteki_farkli_urunler_mukerrer_sanilmaz(client, db_session, kullanici):
    """Stok kodu varsa kimlik odur; aynı açıklama farklı kod = farklı ürün."""
    yukle(client, sablonu_doldur([
        ["OZ-007", "Kablo", "NYA", "Öznur", "2.5", "100", "TL", "Metre", "20"],
        ["OZ-008", "Kablo", "NYA", "Öznur", "2.5", "120", "TL", "Metre", "20"],
    ]))
    assert db_session.query(Product).count() == 2


def test_kesitle_ayrilan_urunler_kodsuz_da_ayri_kalir(client, db_session, kullanici):
    """Stok kodu boşsa kimliğe kesit girer — Öznur'un ham dosyasındaki durum."""
    yukle(client, sablonu_doldur([
        ["", "Kablo", "NYA 450/750", "Öznur", "2.5", "100", "TL", "Metre", "20"],
        ["", "Kablo", "NYA 450/750", "Öznur", "16", "600", "TL", "Metre", "20"],
        ["", "Kablo", "NYA 450/750", "Öznur", "2.5", "999", "TL", "Metre", "20"],
    ]))

    assert db_session.query(Product).count() == 2


# =============================================================================
# Katalog listesi
# =============================================================================


def test_listede_urunun_kendi_para_birimi_gosteriliyor(client, db_session, kullanici):
    """Liste sabit ₺ yazıyordu; Klemsan tamamen EURO veriyor."""
    yukle(client, sablonu_doldur([
        ["KL-002", "Klemens", "REL 24V", "Klemsan", "", "10,58", "EURO", "Adet", "20"],
    ]))

    sayfa = client.get("/catalog").text
    assert "10,58 EUR" in sayfa


def test_listede_kesit_ve_kdv_gorunuyor(client, db_session, kullanici):
    """Kullanıcı içe aktarmanın ne çıkardığını görebilmeli."""
    yukle(client, sablonu_doldur([
        ["OZ-009", "Kablo", "NYA bakır", "Öznur", "3x2.5", "100", "TL", "Metre", "10"],
    ]))

    sayfa = client.get("/catalog").text
    assert "3x2,50 mm²" in sayfa
    assert "KDV %10" in sayfa
    assert "bakır" in sayfa
