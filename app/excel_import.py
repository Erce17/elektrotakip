"""Tedarikçi fiyat listesi okuyucu — saf ayrıştırma katmanı.

Burada DB, ORM, HTTP yok. Girdi bir dosyanın byte'ları, çıktı satır kayıtları ve
ne olduğunu anlatan bir özet.

Neden var: eski içe aktarma sabit sütun sırası bekliyordu (kategori, spec, marka,
fiyat, birim = 0,1,2,3,4), `workbook.active` ile tek sayfa okuyordu ve openpyxl
olduğu için `.xls` açamıyordu. Yani "tedarikçi dosyası aktar" değil "bizim şablonu
doldur" akışıydı. 11 gerçek tedarikçi dosyasında ölçülenler bunu boşa çıkardı:

- **Başlık satırı sabit değil.** 1., 2., 3. veya 5. satırda olabiliyor.
- **Sütun adları çeşitli.** Kod sütunu 7, fiyat sütunu 6 farklı isimle geçiyor.
- **Çoklu sayfa.** Grup Arge 10, Molwex 2, Erse 2 sayfa. Bazı sayfalarda fiyat yok.
- **Bölüm başlıkları veri satırı gibi duruyor** (Kael, Sezgin, Tense, Grup Arge).
- **Para birimi üç türlü geliyor:** ayrı sütun, birim sütununun içinde (`TL/km`),
  ya da birim sütunu diye açılmış ama aslında para birimi tutan sütun (Sezgin).

Ayrıştırıcı tahmin etmez, tanıdığını raporlar: hangi sayfada başlığı nerede buldu,
hangi sütunu neye eşledi, kaç satırı neden atladı. Kullanıcı sonucu görebilsin diye.
"""

import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import openpyxl
import xlrd

# Dosyanın tamamı belleğe okunuyor; fiyat listesi dosyaları küçük olur.
MAX_IMPORT_BYTES = 5 * 1024 * 1024

# Bir sayfada işlenecek azami satır. Kazara açılan devasa dosyada belleği korur.
MAX_SATIR = 50_000

# Başlık kaç satır içinde aranacak. Ölçülen en derini 5. satır (Erse); pay bırakıldı.
BASLIK_ARAMA_DERINLIGI = 15

KM_METRE = Decimal(1000)

# unit_price kolonu Numeric(12,4): 8 tam + 4 ondalık basamak.
# Sınırı aşan değeri kaydetmeye çalışmak Postgres'te DataError'a, yani
# içe aktarmanın tamamının çökmesine yol açıyor.
PRICE_DECIMALS = Decimal("0.0001")
MAX_UNIT_PRICE = Decimal("99999999.9999")


class OkunamadiHatasi(Exception):
    """Dosya hiç açılamadı. Sayfa/sütun bulunamaması bu değil, o özetle raporlanır."""


# =============================================================================
# Fiyat ve birim — eski `catalog.py`'dan taşındı, davranış aynı
# =============================================================================


def parse_price(value) -> Decimal | None:
    """Fiyat girdisini Decimal'e çevirir. Anlaşılmazsa None.

    Tedarikçi listeleri tek bir formatta gelmiyor; ayracın hangisi olduğuna
    şu sırayla karar veriyoruz:

    1. Hem nokta hem virgül varsa → SONDAKİ ondalık ayracıdır.
       '1.200,50' → 1200.50   ·   '1,200.50' → 1200.50
    2. Sadece virgül varsa → ondalık ayracı (Türkçe varsayılan).
       '850,75' → 850.75
    3. Sadece nokta varsa → belirsiz. Birden fazla nokta varsa ya da tek
       noktadan sonra tam 3 hane varsa binlik ayracıdır:
       '1.200' → 1200   ·   '1.234.567' → 1234567   ·   '1200.50' → 1200.50
       Tek istisna: tam sayı kısmı 0 ise ondalıktır ('0.500' → 0.5).

    Decimal kullanılıyor çünkü float ikili tabanda çalışır ve para değerinde
    yuvarlama hatası biriktirir; DB kolonu da Numeric.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        # Excel hücresi zaten sayı; float'ın ikili gösteriminden kaçınmak için
        # metin üzerinden Decimal'e geçiyoruz
        return Decimal(str(value))

    raw = str(value).upper()
    for gereksiz in ("TL", "TRY", "₺", "\xa0"):
        raw = raw.replace(gereksiz, "")
    raw = raw.replace(" ", "").strip()
    if not raw:
        return None

    nokta, virgul = "." in raw, "," in raw
    if nokta and virgul:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")   # 1.200,50
        else:
            raw = raw.replace(",", "")                     # 1,200.50
    elif virgul:
        raw = raw.replace(",", ".")
    elif nokta:
        tam_kisim, _, son_grup = raw.rpartition(".")
        binlik = raw.count(".") > 1 or (len(son_grup) == 3 and tam_kisim.strip("-") != "0")
        if binlik:
            raw = raw.replace(".", "")

    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def quantize_price(value: Decimal) -> Decimal:
    """Fiyatı kolonun tuttuğu basamak sayısına yuvarlar.

    Yuvarlamayı DB'ye bırakmıyoruz: burada yapınca kaydedilen değerle ekranda
    gösterilen değer aynı oluyor.
    """
    return value.quantize(PRICE_DECIMALS, rounding=ROUND_HALF_UP)


# Tedarikçilerin birim sütununa yazdıkları. Soldaki ham yazımlar, sağdaki
# bizim kullandığımız karşılık. Listede olmayan bir yazım olduğu gibi kalır.
BIRIM_KARSILIKLARI = {
    "KM": "KM", "KM.": "KM", "K.M": "KM", "KMTL": "KM", "1000M": "KM",
    "1000MT": "KM", "1000METRE": "KM", "KILOMETRE": "KM",
    "M": "Metre", "MT": "Metre", "MT.": "Metre", "METRE": "Metre", "M.": "Metre",
    "AD": "Adet", "AD.": "Adet", "ADET": "Adet", "ADT": "Adet",
    "KUTU": "Kutu", "KT": "Kutu", "KT.": "Kutu", "BOX": "Kutu",
    "PK": "Paket", "PKT": "Paket", "PAKET": "Paket", "TAKIM": "Takım", "TK": "Takım",
    "KG": "Kg", "KG.": "Kg", "KILO": "Kg",
    "ROLE": "Rulo", "RULO": "Rulo",
}

# Uzunlukla değil sayıyla satılanlar: bunlar hiçbir koşulda 1000'e bölünmez
ADET_BAZLI_BIRIMLER = {"Adet", "Kutu", "Paket", "Takım", "Kg", "Rulo"}


def normalize_unit(raw: str) -> str:
    """Tedarikçinin yazdığı birimi bizim kullandığımız yazıma çevirir.

    'MT', 'mt.', 'metre' hepsi 'Metre' olur. Böylece aynı birim listede üç
    farklı isimle görünmüyor ve KM tespiti güvenilir çalışıyor. Boş girdi boş
    döner — "birim belirtilmemiş" ile "adet" birbirinden ayrılabilsin diye.
    """
    if not raw or not raw.strip():
        return ""
    anahtar = raw.upper().replace(" ", "").replace("/", "").strip()
    return BIRIM_KARSILIKLARI.get(anahtar, raw.strip())


# --- Para birimi -------------------------------------------------------------
# Üç para birimi birden geliyor: Klemsan tamamen EURO, Grup Arge aynı dosyada
# TL ve USD, Viko'da anahtar-priz TL ama LED ve şalt USD.

PARA_KARSILIKLARI = {
    "TL": "TRY", "TRY": "TRY", "₺": "TRY", "TURK LIRASI": "TRY", "T.L.": "TRY",
    "USD": "USD", "$": "USD", "DOLAR": "USD", "USDOLAR": "USD", "AMERIKAN DOLARI": "USD",
    "EUR": "EUR", "EURO": "EUR", "€": "EUR", "AVRO": "EUR",
}


def normalize_currency(raw: str) -> str:
    """Para birimi yazımını ISO koduna çevirir. Tanınmazsa boş döner."""
    if not raw:
        return ""
    return PARA_KARSILIKLARI.get(str(raw).upper().replace(" ", "").strip(), "")


# =============================================================================
# Sütun tanıma
# =============================================================================

# Başlık karşılaştırması için Türkçe harfleri ASCII'ye katlıyoruz: aynı sütun
# 'FİYAT', 'Fiyat', 'FIYAT' diye geçebiliyor ve `str.lower()` 'İ' harfini
# Türkçe kurallarına göre çevirmez.
_HARF_KATLAMA = str.maketrans({
    "İ": "i", "I": "i", "ı": "i", "Ş": "s", "ş": "s", "Ğ": "g", "ğ": "g",
    "Ü": "u", "ü": "u", "Ö": "o", "ö": "o", "Ç": "c", "ç": "c", "Â": "a", "â": "a",
})


def _katla(deger) -> str:
    """Başlığı karşılaştırılabilir hâle getirir: ASCII, küçük harf, tek boşluk."""
    if deger is None:
        return ""
    metin = str(deger).translate(_HARF_KATLAMA).lower()
    metin = metin.replace("\n", " ").replace("\t", " ").replace(".", " ")
    return " ".join(metin.split())


# Ölçülen gerçek başlıklar. Kod sütunu 7, fiyat sütunu 6 farklı isimle geçiyor;
# bu tablo tahmin değil, 11 dosyadan çıkarıldı.
TAM_ESLESMELER = {
    "kod": (
        "stok kodu", "referans", "kod no", "malz kodu", "urun kodu", "siparis kodu",
        "malzeme kodu", "noktasiz stok kodu", "kod", "stok no", "urun no",
    ),
    "ad": (
        "cinsi", "malzeme adi", "urun adi", "stok adi", "stok cinsi", "urun",
        "type", "stok tip", "malzeme cinsi", "ad", "urun ismi", "tanim",
    ),
    "aciklama": (
        "urun aciklamasi", "aciklama", "molwex description", "description",
        "urun aciklama", "ozellik", "teknik ozellik",
    ),
    "grup": (
        "stok grubu", "ana grup", "ana urun gurubu", "stok gurubu", "grup",
        "bolum", "malzeme turu", "kategori", "urun grubu", "ana urun grubu",
    ),
    "marka": ("marka", "brand", "uretici", "firma", "markasi"),
    # Öznur kesiti ürün adında değil ayrı sütunda veriyor; ad 5 satırda birebir
    # aynı olduğu için bu sütun okunmazsa ürünler ayırt edilemiyor.
    "kesit": ("kesit", "kesit (mm2)", "kesit mm2", "kesit (mm²)", "cross section"),
    "birim": ("birim", "olcu birimi", "birimi", "unit"),
    "paket": (
        "paket adet", "kutu miktari", "packing pcs/bag", "kolideki adet",
        "kutudaki adet", "paket", "koli adet", "ambalaj",
    ),
    "fiyat": ("fiyat", "liste fiyat", "liste fiyati", "birim fiyat", "fiyati"),
    "para_birimi": ("para birimi", "doviz", "doviz cinsi", "currency"),
}

# Türkçe ay adları: fiyat sütunu sık sık tarihle adlandırılıyor
# ('2026 ŞUBAT', '2024 KASIM LİSTE FİYATI').
_AYLAR = (
    "ocak", "subat", "mart", "nisan", "mayis", "haziran",
    "temmuz", "agustos", "eylul", "ekim", "kasim", "aralik",
)

# Eşleşme gücü: tam eşleşme, içerik eşleşmesini yener. Viko'da hem 'Fiyat' hem
# 'Tum Fiyatlar', Öznur'da hem 'Fiyat' hem 'Fiyat (ham)' var — doğrusu kısa olan.
TAM = 2
ICERIK = 1


def _kelime_dizisi_var(kelimeler: list[str], aranan: list[str]) -> bool:
    """`aranan` kelime dizisi `kelimeler` içinde bitişik olarak geçiyor mu."""
    if not aranan or len(aranan) > len(kelimeler):
        return False
    return any(
        kelimeler[i:i + len(aranan)] == aranan
        for i in range(len(kelimeler) - len(aranan) + 1)
    )


def sutun_tipi(baslik) -> tuple[str, int] | None:
    """Bir başlık hücresini kanonik alana eşler. Eşleşmezse None.

    Sıra önemli: para birimi önce bakılır. Klemsan'ın 'Fiyat Listesi Döviz Cinsi'
    başlığı 'fiyat' kelimesini içeriyor ama fiyat sütunu değil.
    """
    k = _katla(baslik)
    if not k:
        return None

    # 1. Para birimi — 'fiyat' içerse bile önce bu kazanır.
    if k in TAM_ESLESMELER["para_birimi"]:
        return ("para_birimi", TAM)
    if any(ip in k for ip in ("para birimi", "doviz", "currency")):
        return ("para_birimi", ICERIK)

    # 2. Tam eşleşme.
    for alan, adaylar in TAM_ESLESMELER.items():
        if k in adaylar:
            return (alan, TAM)

    # 3. Fiyat: ya 'fiyat' geçiyordur ya da tarihle adlandırılmıştır.
    if "fiyat" in k:
        return ("fiyat", ICERIK)
    if any(ay in k.split() for ay in _AYLAR) and any(p.isdigit() for p in k.split()):
        return ("fiyat", ICERIK)

    # 4. Kalanlar için gevşek eşleşme — ama **kelime bazlı**, alt dize değil.
    # Alt dize aramak felaket veriyordu: 'ADET' içinde "ad", 'Turuncu' içinde
    # "urun", 'KODLAMA MALZEMESİ' içinde "kod" geçiyor. Molwex'te 699 ürün satırı
    # bu yüzden başlık sanılıp atılmıştı.
    kelimeler = k.split()
    for alan in ("kod", "aciklama", "ad", "grup", "marka", "kesit", "paket", "birim"):
        for aday in TAM_ESLESMELER[alan]:
            if _kelime_dizisi_var(kelimeler, aday.split()):
                return (alan, ICERIK)
    return None


def sutunlari_esle(baslik_satiri: list) -> dict[str, int]:
    """Başlık satırından `{kanonik alan: sütun indeksi}` çıkarır.

    Aynı alana birden çok aday çıkarsa tam eşleşen kazanır, eşitlikte soldaki.
    """
    en_iyi: dict[str, tuple[int, int]] = {}  # alan -> (guc, indeks)
    for indeks, hucre in enumerate(baslik_satiri):
        eslesme = sutun_tipi(hucre)
        if eslesme is None:
            continue
        alan, guc = eslesme
        mevcut = en_iyi.get(alan)
        if mevcut is None or guc > mevcut[0]:
            en_iyi[alan] = (guc, indeks)
    return {alan: indeks for alan, (_, indeks) in en_iyi.items()}


def baslik_satirini_bul(satirlar: list[list]) -> tuple[int, dict[str, int]] | None:
    """Sayfanın ilk satırlarını tarayıp başlık satırını bulur.

    Ölçüm: başlık 1. (Grup Arge, Klemsan, Molwex, Öznur, Pofaco, Viko),
    2. (Sezgin, Tense), 3. (Kael) veya 5. (Erse) satırda olabiliyor. Sabit
    varsayım yerine puanlama: en çok sütunu tanınan satır kazanır.

    Fiyat sütunu şart. Fiyatsız sayfa (Molwex 'Tüm Kodlar') kataloğa giremez;
    onu başlık bulamamış saymak, yarım veri almaktan iyidir.
    """
    en_iyi: tuple[int, int, dict[str, int]] | None = None  # (puan, indeks, esleme)
    for indeks, satir in enumerate(satirlar[:BASLIK_ARAMA_DERINLIGI]):
        esleme = sutunlari_esle(satir)
        if "fiyat" not in esleme or len(esleme) < 2:
            continue
        puan = len(esleme)
        if en_iyi is None or puan > en_iyi[0]:
            en_iyi = (puan, indeks, esleme)
    if en_iyi is None:
        return None
    return (en_iyi[1], en_iyi[2])


# =============================================================================
# Dosya okuma — .xlsx ve .xls tek arayüzde
# =============================================================================


def _xlsx_sayfalari(icerik: bytes):
    try:
        workbook = openpyxl.load_workbook(filename=io.BytesIO(icerik), data_only=True,
                                          read_only=True)
    except Exception as hata:
        # openpyxl bozuk/şifreli/eski format dosyalarda çok çeşitli hata atar;
        # kullanıcı için hepsi aynı sonuca çıkıyor.
        raise OkunamadiHatasi(str(hata)) from hata
    try:
        for ad in workbook.sheetnames:
            sayfa = workbook[ad]
            satirlar = []
            for sira, satir in enumerate(sayfa.iter_rows(values_only=True)):
                if sira >= MAX_SATIR:
                    break
                satirlar.append(list(satir))
            yield ad, satirlar
    finally:
        workbook.close()


def _xls_sayfalari(icerik: bytes):
    try:
        kitap = xlrd.open_workbook(file_contents=icerik)
    except Exception as hata:
        raise OkunamadiHatasi(str(hata)) from hata
    for sayfa in kitap.sheets():
        satirlar = [
            [h.value for h in sayfa.row(r)]
            for r in range(min(sayfa.nrows, MAX_SATIR))
        ]
        yield sayfa.name, satirlar


def sayfalari_oku(icerik: bytes, dosya_adi: str):
    """Dosyayı biçimine göre açar, `(sayfa adı, satırlar)` üretir.

    `.xls` ayrı bir kütüphane ister (openpyxl yalnızca OOXML okur). Sahada eski
    format hâlâ yaygın; tedarikçiden gelen dosyayı reddetmek kötü karşılanıyor.
    """
    if dosya_adi.lower().endswith(".xls"):
        return _xls_sayfalari(icerik)
    return _xlsx_sayfalari(icerik)


# =============================================================================
# Satır ayrıştırma
# =============================================================================


@dataclass(frozen=True)
class SatirKaydi:
    """Bir tedarikçi satırının ayrıştırılmış hâli. Henüz ürün değil."""

    sayfa: str
    satir_no: int          # dosyadaki 1 tabanlı satır numarası; hata mesajı için
    kod: str = ""
    ad: str = ""
    aciklama: str = ""
    grup: str = ""
    marka: str = ""
    kesit: str = ""        # dosyanın kendi kesit sütunu; metinden okunana tercih edilir
    fiyat: Decimal | None = None   # None = hücre doluydu ama okunamadı
    para_birimi: str = ""
    birim: str = ""
    paket: str = ""


@dataclass
class SayfaOzeti:
    """Bir sayfada ne olduğu. Kullanıcıya 'neden bu kadar az ürün geldi' cevabı."""

    ad: str
    baslik_satiri: int | None = None       # 1 tabanlı, None = başlık bulunamadı
    sutunlar: dict[str, int] = field(default_factory=dict)
    okunan: int = 0
    bolum_basligi: int = 0                 # fiyat hücresi boş olan satırlar
    tekrar_eden_baslik: int = 0
    kimliksiz: int = 0                     # kodu da adı da açıklaması da boş
    fiyat_okunamadi: int = 0               # 'Bilgi Alınız', 'FİYAT SORUNUZ'


@dataclass
class DosyaOzeti:
    satirlar: list[SatirKaydi] = field(default_factory=list)
    sayfalar: list[SayfaOzeti] = field(default_factory=list)

    @property
    def atlanan_sayfalar(self) -> list[str]:
        return [s.ad for s in self.sayfalar if s.baslik_satiri is None]


def _metin(satir: list, indeks: int | None) -> str:
    if indeks is None or indeks >= len(satir):
        return ""
    deger = satir[indeks]
    if deger is None:
        return ""
    # xlrd sayısal hücreleri float verir: '001 002' gibi kodlar bozulmasın diye
    # tam sayı olanları noktasız yazıyoruz.
    if isinstance(deger, float) and deger.is_integer():
        return str(int(deger))
    return str(deger).strip()


def _bolum_basligi_mi(satir: list, sutunlar: dict[str, int]) -> bool:
    """Bölüm başlığı mı, ürün mü?

    Ayırt edici kural: **fiyat hücresi tamamen boşsa bölüm başlığıdır.** Kael'de
    'DİNAMİK KOMPANZASYON', Sezgin'de 'TOR', Tense'de 'STATİK VAR GENERATOR'
    böyle geçiyor. Buna karşılık Tense'in 'Bilgi Alınız' yazan satırları gerçek
    ürün — hücre dolu, sadece sayı değil. Bu yüzden "boş" ile "okunamadı" ayrı.

    İkinci kural: ad/açıklama sütunları tanınmış, hepsi boş **ve** fiyat hücresi
    sayıya çevrilemiyorsa başlıktır. Grup Arge'de bölüm satırı fiyat sütununa
    '2026-1' yazıyor ve ürün adı boş — ikisi birden.

    ⚠️ İkinci kuralda "fiyat sayıya çevrilemiyor" şartı olmazsa gerçek ürün
    kaybedilir: Viko'da 822 satırın ürün adı boş ama fiyatı geçerli (ad bilgisi
    başka sütunda). Ölçüldü, tahmin değil.
    """
    fiyat_hucresi = _metin(satir, sutunlar.get("fiyat"))
    if not fiyat_hucresi:
        return True
    ad_sutunlari = [sutunlar[a] for a in ("ad", "aciklama") if a in sutunlar]
    if ad_sutunlari and not any(_metin(satir, i) for i in ad_sutunlari):
        return parse_price(fiyat_hucresi) is None
    return False


def _tekrar_eden_baslik_mi(satir: list, sutunlar: dict[str, int]) -> bool:
    """Sayfa ortasında yeniden yazılmış başlık satırı (Erse 'Baskı Formatı')."""
    taninan = sum(1 for i in sutunlar.values() if sutun_tipi(_metin(satir, i)) is not None)
    return taninan >= 2


def _para_ve_birim(satir: list, sutunlar: dict[str, int]) -> tuple[str, str]:
    """Para birimi ve birimi çözer. Üçü de gerçek dosyalardan çıkmış durum:

    1. Ayrı `PARA BİRİMİ` sütunu (Grup Arge, Molwex, Klemsan).
    2. Birim sütununun içinde: Öznur `TL/km`, Erse `t/m`.
    3. `BİRİM` diye açılmış ama para birimi tutan sütun (Sezgin: `USD`).
    """
    para = normalize_currency(_metin(satir, sutunlar.get("para_birimi")))
    ham_birim = _metin(satir, sutunlar.get("birim"))

    if "/" in ham_birim:
        sol, _, sag = ham_birim.partition("/")
        cevrilen = normalize_currency(sol)
        if cevrilen:
            para = para or cevrilen
            ham_birim = sag

    # Sütun adı 'BİRİM' ama içerik para birimi: birim bilgisi değil kur bilgisi.
    if not para:
        cevrilen = normalize_currency(ham_birim)
        if cevrilen:
            return cevrilen, ""

    return para, normalize_unit(ham_birim)


def sayfayi_ayristir(sayfa_adi: str, satirlar: list[list]) -> tuple[SayfaOzeti, list[SatirKaydi]]:
    ozet = SayfaOzeti(ad=sayfa_adi)
    bulunan = baslik_satirini_bul(satirlar)
    if bulunan is None:
        return ozet, []

    baslik_indeksi, sutunlar = bulunan
    ozet.baslik_satiri = baslik_indeksi + 1
    ozet.sutunlar = sutunlar

    kayitlar: list[SatirKaydi] = []
    for indeks in range(baslik_indeksi + 1, len(satirlar)):
        satir = satirlar[indeks]
        if not any(h not in (None, "") for h in satir):
            continue
        if _tekrar_eden_baslik_mi(satir, sutunlar):
            ozet.tekrar_eden_baslik += 1
            continue
        if _bolum_basligi_mi(satir, sutunlar):
            ozet.bolum_basligi += 1
            continue

        kod = _metin(satir, sutunlar.get("kod"))
        ad = _metin(satir, sutunlar.get("ad"))
        aciklama = _metin(satir, sutunlar.get("aciklama"))
        if not (kod or ad or aciklama):
            # Fiyatı olan ama hiçbir kimliği olmayan satır ürün olamaz. Viko'nun
            # PDF çevriminde böyle artıklar var: satırda yalnız tutar kalmış.
            ozet.kimliksiz += 1
            continue

        fiyat = parse_price(satir[sutunlar["fiyat"]] if sutunlar["fiyat"] < len(satir) else None)
        if fiyat is None:
            # 'Bilgi Alınız' (Tense), 'FİYAT SORUNUZ', 'KDV DAHİL NET FİYAT'
            # (Grup Arge). Ürün gerçek, fiyatı sonra girilecek.
            ozet.fiyat_okunamadi += 1

        para, birim = _para_ve_birim(satir, sutunlar)
        kayitlar.append(
            SatirKaydi(
                sayfa=sayfa_adi,
                satir_no=indeks + 1,
                kod=kod,
                ad=ad,
                aciklama=aciklama,
                grup=_metin(satir, sutunlar.get("grup")),
                marka=_metin(satir, sutunlar.get("marka")),
                kesit=_metin(satir, sutunlar.get("kesit")),
                fiyat=fiyat,
                para_birimi=para,
                birim=birim,
                paket=_metin(satir, sutunlar.get("paket")),
            )
        )
        ozet.okunan += 1

    return ozet, kayitlar


# Excel'in kendi verdiği sayfa adları. Bunlar kategori adı olarak işe yaramaz.
GENEL_SAYFA_ADLARI = {"sheet", "sheet1", "sayfa1", "sayfa", "worksheet", "tablo1"}


def sayfa_adi_kategori_olur_mu(ad: str) -> bool:
    """Sayfa adı kategori olarak kullanılabilir mi.

    Grup sütunu olmayan dosyalarda (Öznur, Pofaco) sayfa adı tek anlamlı kategori
    ipucu. Ama 'Sayfa1' gibi Excel'in kendi verdiği adlar bilgi taşımaz.
    """
    return _katla(ad) not in GENEL_SAYFA_ADLARI


def oku(icerik: bytes, dosya_adi: str) -> DosyaOzeti:
    """Dosyanın tamamını ayrıştırır. Tüm sayfalar işlenir.

    Fiyat sütunu bulunamayan sayfa atlanır ve özette adıyla raporlanır — Molwex'in
    'Tüm Kodlar' sayfası böyle: gerçek bir sayfa ama fiyat tutmuyor.
    """
    sonuc = DosyaOzeti()
    for sayfa_adi, satirlar in sayfalari_oku(icerik, dosya_adi):
        ozet, kayitlar = sayfayi_ayristir(sayfa_adi, satirlar)
        sonuc.sayfalar.append(ozet)
        sonuc.satirlar.extend(kayitlar)
    return sonuc
